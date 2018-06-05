# coding=utf-8
#
import unittest
import logging
import logging.config
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time

import os
from openpyxl import load_workbook

logging.config.fileConfig('logging.conf')
logger = logging.getLogger('main')

class Perform:
   def test(sendmes):
      logger.info(sendmes['en_user'])

      return 'ture'
   def doLoginSession(sendmes):
          try:
               logger.info('----- 开始----')
               driver = webdriver.Chrome(executable_path="bin/chromedriver.exe")

               logger.info('-----Install driver----')
               driver.implicitly_wait(3)
               driver.get("http://113.231.68.122/seeyon/index.jsp")
               elemuser = driver.find_element_by_id("login_username")
               elemuser.send_keys(sendmes['en_user'])
               elempswd = driver.find_element_by_id("login_password")
               elempswd.send_keys(sendmes['en_passwd'])
               elempswd.send_keys(Keys.RETURN)
               driver.get("http://113.231.68.122/seeyon/collaboration/collaboration.do?method=newColl&templateId=-6727894578816138688")
          except Exception as e:
               return 'chromeerror'
          try:
               logger.info("--执行填写--")

               getFrame = driver.find_element_by_id("zwIframe")

               driver.switch_to.frame(getFrame)
               datakey=driver.find_element_by_xpath('// *[ @ id = "field0008"]')
               for i in range(20):
                  datakey.send_keys(Keys.BACKSPACE)
               driver.find_element_by_xpath('// *[ @ id = "field0008"]').send_keys(sendmes['lastday'])

               logger.info('外部')

               driver.execute_script('''$('#field0056_span :last-child').click()''')

               logger.info('click')

               logger.info('ok')

               driver.switch_to.default_content()
               getFrame_checkfrom = driver.find_element_by_xpath('//*[contains(@id,"main_iframe_content")]')
               driver.switch_to.frame(getFrame_checkfrom)
               lenth=0
               indexarry = []
               indexpersent =[]
               #
               filename='Mdict/'+sendmes['send_role']+'.txt'
               logger.info(filename)
               if os.path.exists(filename) and os.path.getsize(filename)!= 0:
                   with open(filename, 'r') as f:
                        for line in f.readlines():
                           indexarry.append(line.strip().split('#')[1])
                           indexpersent.append(line.strip().split('#')[2])
               else:
                  return 'no temp file found'
               #
               logger.info(indexarry)
               logger.info(indexpersent)
               driver.find_element_by_xpath('//*[@id = "rpInputChange"]').send_keys(Keys.BACKSPACE)
               driver.find_element_by_xpath('//*[@id = "rpInputChange"]').send_keys('5')
               driver.find_element_by_xpath('//*[@id="grid_go"]').click()

               #driver.implicitly_wait(5)
               #WebDriverWait(driver, 超时时长, 调用频率, 忽略异常).until(可执行方法, 超时时返回的信息)

               loder=(By.XPATH,'//*[@id="row7299443944487013551"]/td[1]/div/input')
               WebDriverWait(driver,20,0.5).until(EC.presence_of_element_located(loder))

               for index in indexarry:
                   mypath= '//*[@id="row'+index+'"]/td[1]/div/input'
                   selectIndex = driver.find_element_by_xpath(mypath)
                   selectIndex.click()

               driver.switch_to.default_content()
               donefrom = driver.find_element_by_link_text("确定")
               donefrom.click()
               #-----------------填写
               getFrame = driver.find_element_by_id("zwIframe")
               logger.info('外部框')
               logger.info(getFrame)
               driver.switch_to.frame(getFrame)
               time.sleep(1)
               #填写占比
               block1s= driver.find_elements_by_xpath('//*[@id="field0059_txt"]')
               logger.info(block1s)
               for block in range(len(block1s)):
                     block1s[block].send_keys(indexpersent[block])

               showAdd = driver.find_element_by_xpath('//*[@id="field0065"]')
               showAdd.click()
               myadd = driver.find_element_by_xpath('//*[@id="addEmptyImg"]')
               path=sendmes['xml_path']
               wb = load_workbook(path)

               logger.info(wb.sheetnames)
               logger.info(sendmes['group'])
               logger.info(sendmes['group'])
               sheetname=sendmes['group']
               sheet = wb.get_sheet_by_name(sheetname)
               tabRow = sheet.max_row
               
               logger.info('num:'+str(tabRow))
               for i in range(2, tabRow):
                  myadd.click()

               for i in range(1, tabRow):
                     logger.info(i)
                     sheetB=driver.find_element_by_xpath('//*[@id="formson_24252"]/tbody/tr['+str(i)+']/td[2]/div/span/textarea')
                     sheetB.send_keys(sheet['B' + str(i+1)].value)
                     sheetTime=driver.find_element_by_xpath('//*[@id="formson_24252"]/tbody/tr['+str(i)+']/td[3]/div/span/textarea')
                     sheetTime.send_keys(sendmes['xml_lastday'])
                     sheetC=driver.find_element_by_xpath('//*[@id="formson_24252"]/tbody/tr['+str(i)+']/td[4]/div/span/textarea')
                     sheetC.send_keys(sheet['C'+str(i+1)].value)
                     sheetD=driver.find_element_by_xpath('//*[@id="formson_24252"]/tbody/tr['+str(i)+']/td[5]/div/span/input[2]')
                     sheetD.send_keys(sheet['D'+str(i+1)].value)

                     #sheetE=driver.find_element_by_xpath('//*[@id="formson_24252"]/tbody/tr['+str(i)+']/td[6]/div/span/textarea')
                     #sheetE.send_keys(sheet['E'+str(i+1)].value)
               logger.info("return to main")
               driver.switch_to.default_content()
               logger.info("next do save")
               driver.find_element_by_id("saveDraft_a").click()
               logger.info('-----------------以存储至待发-----------------------')
               logger.info("以存储至待发")
               driver.quit()
               return 'ture'
          except Exception as e:
               logger.info(e)
               logger.error(e)
               driver.quit()
               return 'false'
